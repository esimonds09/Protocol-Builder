import json
import os
import pandas as pd
import ipywidgets as ipw
from openpyxl import load_workbook
from IPython.display import display
from config import config, Headers, FixedHiPrBindCalcs


class ExcelData:
    def __init__(self):
        self.file_path = f"templates/Inventory Tracking.xlsx"
        self.worksheet = 'Project Specific Reagents'
        self.reagent_data = self.import_data()

    def import_data(self):
        reagent_data = pd.read_excel(self.file_path, sheet_name=self.worksheet, index_col=[0], skiprows=1)
        return reagent_data

    def get_projects(self):
        project_list = list(self.reagent_data.index.unique())
        project_dict = {value: project_list.index(value) + 1 for value in project_list}
        scheme_dict = {}
        for project, value in project_dict.items():
            scheme_dict[value] = [int(num) for num in list(self.reagent_data.loc[project]["Scheme"].unique()) if num > 0]
        project_dict["..."] = 0
        scheme_dict[0] = []
        return project_dict, scheme_dict

    def get_scheme(self, project):
        reagent_data = self.reagent_data.loc[project]
        scheme_list = list(reagent_data[~reagent_data["Reagent"].str.contains('standard')]["Scheme"])
        return scheme_list

    def get_reagent_data(self, project, scheme):
        reagent_data = self.reagent_data.loc[project].sort_values("Assay")
        reagent_data_wo_standard = reagent_data[~reagent_data["Reagent"].str.contains('standard', case=False)]
        reagent_data_scheme = reagent_data_wo_standard.loc[reagent_data_wo_standard["Scheme"].isin([0, scheme])]
        return reagent_data_scheme

    def get_standard_data(self, project):
        reagent_data = self.reagent_data.loc[project]
        reagent_data_w_standard = reagent_data[reagent_data["Reagent"].str.contains('standard', case=False)]
        return reagent_data_w_standard


class ProtocolForm:
    def __init__(self):
        self.proj_details = ProjectDetails()
        self.plate_details = PlateDetails()
        self.dil_details = DilutionDetails()
        self.st_details = StandardDetails()
        self.headers = Headers()
        self.data_dict = {}
        self.display_message = ipw.VBox()
        self.run_notes = ipw.Textarea(
            description="Run notes:",
            style=self.headers.style
        )
        self.capture_inputs_button = ipw.Button(
            description="Capture Inputs",
            button_style="info"
        )
        self.capture_inputs_button.on_click(self.capture_inputs)

        self.add_standard_button = ipw.Button(
            description="Add Standard",
            button_style="info"
        )
        self.add_standard_button.on_click(self.add_standard_data)
        self.add_standard_clicked = False

        self.project_details = self.proj_details.proj_details()
        self.pl_details = self.plate_details.plate_details()
        self.dilution_details = self.dil_details.dilution_details()
        self.standard_details = ipw.VBox()

        self.display_form()

    def display_form(self):
        input_boxes = ipw.VBox([
            self.project_details,
            self.pl_details,
            self.dilution_details,
            self.standard_details
        ])

        form_display = ipw.VBox([
            ipw.HTML("<h3>Protocol Form</h3>"),
            input_boxes,
            self.add_standard_button,
            ipw.HTML("<h5><b>Execution Notes:</b></h5>"),
            self.run_notes,
            ipw.HTML("<br>"),
            self.capture_inputs_button,
            self.display_message,
            ])

        display(form_display)

    def add_standard_data(self, event):
        project = self.proj_details.project_choice.label
        if self.add_standard_clicked:
            self.standard_details.children = []
            self.add_standard_button.description = "Add Standard"
            self.add_standard_clicked = False
        else:
            if project == "...":
                self.standard_details.children = [ipw.HTML("CHOOSE PROJECT")]
            else:
                self.standard_details.children = [self.st_details.standard_details(project)]
                self.add_standard_button.description = "Hide Standard"
                self.add_standard_clicked = True

    def capture_inputs(self, event):
        project = self.proj_details.project_choice.label
        self.display_message.children = [ipw.HTML("""<b>Data captured. Please click 'Display Outputs' button below.<br>
                                                 If inputs are updated, remember to click both buttons again.</b>""")]
        self.data_dict.update(self.proj_details.capture_inputs())
        self.data_dict.update(self.plate_details.capture_inputs())
        self.data_dict.update(self.dil_details.capture_inputs())
        self.data_dict.update(self.st_details.capture_inputs())
        self.data_dict["run_notes"] = self.run_notes.value

        with open('captured_data.json', 'w') as captured_data:
            json.dump(self.data_dict, captured_data, indent=4)


class ProjectDetails:
    def __init__(self):
        self.headers = Headers()
        # temp project dict from excel
        projects, schemes = ExcelData().get_projects()
        self.project_header = ipw.HTML("<h5><b>Project Details:</b></h5>")
        self.project_choice = ipw.Dropdown(
            description="Choose project: ",
            options=projects,
            value=projects['...'],
            style=self.headers.style
        )
        self.project_id = ipw.Text(
            description="Enter project name: ",
            style=self.headers.style
        )
        self.project_file_option = ipw.Text(
            description="File tag (optional): ",
            style=self.headers.style
        )
        self.project_type = ipw.Dropdown(
            description="Choose run type: ",
            options=["SSF_LS", "SSF_LR", "SSF_LR_DSS", "SSF_DEV", "Fermentation"],
            style=self.headers.style
        )
        self.project_scheme = ipw.Dropdown(
            description="Choose a scheme: ",
            options=schemes[self.project_choice.value],
            style=self.headers.style
        )

        def update_scheme(project):
            scheme_options = schemes[project]
            self.project_scheme.options = scheme_options
        ipw.interactive(update_scheme, project=self.project_choice)

    def proj_details(self):
        proj_details = ipw.VBox([
            self.project_header,
            ipw.HBox([
                self.project_choice,
                self.project_scheme
            ]),
            ipw.HBox([
                self.project_id,
                self.project_file_option
            ]),
            self.project_type
        ])

        return proj_details

    def capture_inputs(self):
        input_dict = dict(
            project=self.project_choice.label,
            project_name_id=self.project_choice.value,
            project_scheme=self.project_scheme.value,
            proj_id=self.project_id.value,
            proj_file_option=self.project_file_option.value,
            proj_type=self.project_type.value
        )
        return input_dict


class PlateDetails:
    def __init__(self):
        self.headers = Headers()
        self.plate_header = ipw.HTML('<h5><b>Plate Details:</b></h5>')
        self.total_source = ipw.IntText(
            description="Source Plates: ",
            style=self.headers.style
        )
        self.replicates = ipw.Dropdown(
            description='Replicates: ',
            options=['0', 'n + 1', 'n + 2', 'n * 2'],
            style=self.headers.style
        )
        self.total_predilution = 0
        self.include_pd = ipw.Checkbox(
            description="Predilution Plates: ",
            style=self.headers.style,
            indent=False
        )
        self.pd_vol_head = ipw.HTML('<h5><b>Predilution DBI vol (uL):</b></h5>')
        self.pd_1_vol = ipw.IntText(description='Plate 1: ', style=self.headers.style)
        self.pd_2_vol = ipw.IntText(description='Plate 2: ', style=self.headers.style)
        self.pd_3_vol = ipw.IntText(description='Plate 3: ', style=self.headers.style)
        self.pd_4_vol = ipw.IntText(description='Plate 4: ', style=self.headers.style)
        self.pd_vol_display = ipw.VBox([
            self.pd_vol_head,
            self.pd_1_vol,
            self.pd_2_vol,
            self.pd_3_vol,
            self.pd_4_vol
        ])
        self.pd_spike_head = ipw.HTML('<h5><b>Predilution Spike vol (uL):</b></h5>')
        self.pd_1_spike = ipw.IntText(description='Spike 1: ', style=self.headers.style)
        self.pd_2_spike = ipw.IntText(description='Spike 2: ', style=self.headers.style)
        self.pd_3_spike = ipw.IntText(description='Spike 3: ', style=self.headers.style)
        self.pd_4_spike = ipw.IntText(description='Spike 4: ', style=self.headers.style)
        self.pd_spike_display = ipw.VBox([
            self.pd_spike_head,
            self.pd_1_spike,
            self.pd_2_spike,
            self.pd_3_spike,
            self.pd_4_spike
        ])
        self.predilution_display = ipw.HBox()

        def get_predilution(pd_check):
            if pd_check:
                self.predilution_display.children = [self.pd_vol_display, self.pd_spike_display]
            else:
                self.predilution_display.children = []

        ipw.interactive(get_predilution, pd_check=self.include_pd)

    def plate_details(self):
        plate_details = ipw.VBox([
            self.plate_header,
            ipw.HBox([
                self.total_source,
                self.replicates,
            ]),
            self.include_pd,
            self.predilution_display
        ])
        return plate_details

    def capture_inputs(self):
        if self.include_pd:
            self.total_predilution = 4 - [
                self.pd_1_vol.value,
                self.pd_2_vol.value,
                self.pd_3_vol.value,
                self.pd_4_vol.value
            ].count(0)

            plate_dict = dict(
                source=self.total_source.value,
                replicates=self.replicates.value,
                pd=self.total_predilution,
                pd_vols=dict(
                    pd_1_vol=self.pd_1_vol.value,
                    pd_2_vol=self.pd_2_vol.value,
                    pd_3_vol=self.pd_3_vol.value,
                    pd_4_vol=self.pd_4_vol.value,
                ),
                pd_spikes=dict(
                    pd_1_spike=self.pd_1_spike.value,
                    pd_2_spike=self.pd_2_spike.value,
                    pd_3_spike=self.pd_3_spike.value,
                    pd_4_spike=self.pd_4_spike.value
                ),
            )
            return plate_dict


class DilutionDetails:
    def __init__(self):
        self.headers = Headers()
        self.db_head = ipw.HTML('<h5><b>Source Resuspension / Greiner Dilution Volumes: </b></h5>')
        self.dbi_vol = ipw.IntText(
            description="DBI uL required / well: ",
            style=self.headers.style
        )
        self.dbii_vol = ipw.IntText(
            description="DBII uL required / well: ",
            style=self.headers.style
        )
        self.input_vol_head = ipw.HTML('<h5><b>Serial Dilutions: </b></h5>')
        self.input_vol_1 = ipw.IntText(description="Volume 1 (uL): ", style=self.headers.style)
        self.input_vol_2 = ipw.IntText(description="Volume 2 (uL): ", style=self.headers.style)
        self.input_vol_3 = ipw.IntText(description="Volume 3 (uL): ", style=self.headers.style)
        self.input_vol_4 = ipw.IntText(description="Volume 4 (uL): ", style=self.headers.style)
        self.point_dilution = ipw.Dropdown(
            description="Choose 4 or 8 point dilution: ",
            options={'4 point': 4, '8 point': 8},
            value=4,
            style=self.headers.style
        )
        self.cell_pellet = ipw.IntText(
            description="Cell pellet size:",
            value=0,
            style=self.headers.style
        )

    def dilution_details(self):
        input_vol_display = ipw.VBox([
            self.db_head,
            ipw.HBox([
                self.dbi_vol,
                self.dbii_vol
            ]),
            self.input_vol_head,
            ipw.HBox([
                self.point_dilution,
                self.cell_pellet,
            ]),
            self.input_vol_1,
            self.input_vol_2,
            self.input_vol_3,
            self.input_vol_4
        ])
        return input_vol_display

    def capture_inputs(self):
        dilution_dict = dict(
            dbi_vol=self.dbi_vol.value,
            dbii_vol=self.dbii_vol.value,
            dil_vols=dict(
                dil_vol_1=self.input_vol_1.value,
                dil_vol_2=self.input_vol_2.value,
                dil_vol_3=self.input_vol_3.value,
                dil_vol_4=self.input_vol_4.value
            ),
            points=self.point_dilution.value,
            cell_resus=self.cell_pellet.value
        )
        return dilution_dict


class StandardDetails:
    def __init__(self):
        self.headers = Headers()
        self.standard_stock_header = ipw.HTML('<h5><b>Standard Details</b></h5>')
        self.standard_stock_conc = ipw.FloatText(
            description='Conc. (ug/uL)',
            value=0
        )
        self.standard_stock_mw = ipw.IntText(
            description='MW (g/mol)',
            value=0
        )

        self.standard_box_header = ipw.HTML('<h5><b>Standard Concentrations to be used:</b></h5>')
        self.standard_conc_1 = ipw.FloatText(description='#1 nM: ')
        self.standard_conc_2 = ipw.FloatText(description='#2 nM: ', disabled=True)
        self.standard_conc_3 = ipw.FloatText(description='#3 nM: ', disabled=True)
        self.standard_conc_4 = ipw.FloatText(description='#4 nM: ', disabled=True)
        self.standard_conc_5 = ipw.FloatText(description='#5 nM: ', disabled=True)
        self.standard_conc_6 = ipw.FloatText(description='#6 nM: ', disabled=True)
        self.standard_fold_1 = ipw.FloatText(description='Fold #1: ')
        self.standard_fold_2 = ipw.FloatText(description='Fold #2: ')
        self.standard_fold_3 = ipw.FloatText(description='Fold #3: ')
        self.standard_fold_4 = ipw.FloatText(description='Fold #4: ')
        self.standard_fold_5 = ipw.FloatText(description='Fold #5: ')
        self.standard_wells = ipw.IntText(
            description='# of replicates for standard: ',
            value=2,
            style=self.headers.style
        )
        self.standard_plates = ipw.IntText(
            description='# of source plates with standard: ',
            style=self.headers.style
        )
        self.standard_vol = ipw.IntText(
            description='Volume uL per well: ',
            style=self.headers.style
        )
        self.calc_nm_button = ipw.Button(
            description="Calculate nM",
            button_style="info"
        )
        self.calc_nm_button.on_click(self.calculate_nm)
        self.calc_error_msg = ipw.VBox()

        self.standard_info = ipw.VBox([
            self.standard_stock_header,
            ipw.HBox([
                self.standard_stock_conc,
                self.standard_stock_mw
            ]),
            self.standard_box_header,
            ipw.HBox([
                ipw.VBox([
                    self.standard_conc_1,
                    self.standard_conc_2,
                    self.standard_conc_3,
                    self.standard_conc_4,
                    self.standard_conc_5,
                    self.standard_conc_6
                ]),
                ipw.VBox([
                    self.standard_fold_1,
                    self.standard_fold_2,
                    self.standard_fold_3,
                    self.standard_fold_4,
                    self.standard_fold_5
                ]),
                ipw.VBox([
                    self.standard_wells,
                    self.standard_plates,
                    self.standard_vol
                ])
            ]),
            self.calc_nm_button
            ])

    def calculate_nm(self, event):
        try:
            self.standard_conc_2.value = round((self.standard_conc_1.value / self.standard_fold_1.value), 4)
            self.standard_conc_3.value = round((self.standard_conc_2.value / self.standard_fold_2.value), 4)
            self.standard_conc_4.value = round((self.standard_conc_3.value / self.standard_fold_3.value), 4)
            self.standard_conc_5.value = round((self.standard_conc_4.value / self.standard_fold_4.value), 4)
            self.standard_conc_6.value = round((self.standard_conc_5.value / self.standard_fold_5.value), 4)
        except ZeroDivisionError:
            self.calc_error_msg.children += (ipw.HTML("One of the folds is zero!"),)

    def standard_details(self, project):
        self.update_standard_stock(project)
        standard_display = ipw.VBox([
            self.standard_info,
            self.calc_error_msg
        ])
        return standard_display

    def update_standard_stock(self, project):
        # code block for excel data
        standard_data = ExcelData().get_standard_data(project)
        self.standard_stock_conc.value = standard_data['Conc. (ug/uL)'].squeeze()
        self.standard_stock_mw.value = standard_data['MW (g/mol)'].squeeze()

    def capture_inputs(self):
        standard_dict = dict(
            standard_stock=dict(
                standard_stock_conc=self.standard_stock_conc.value,
                standard_stock_mw=self.standard_stock_mw.value
            ),
            standard_concs=dict(
                standard_conc_1=self.standard_conc_1.value,
                standard_conc_2=self.standard_conc_2.value,
                standard_conc_3=self.standard_conc_3.value,
                standard_conc_4=self.standard_conc_4.value,
                standard_conc_5=self.standard_conc_5.value,
                standard_conc_6=self.standard_conc_6.value,
            ),
            standard_folds=dict(
                standard_fold_1=self.standard_fold_1.value,
                standard_fold_2=self.standard_fold_2.value,
                standard_fold_3=self.standard_fold_3.value,
                standard_fold_4=self.standard_fold_4.value,
                standard_fold_5=self.standard_fold_5.value,
            ),
            standard_wells=self.standard_wells.value,
            standard_plates=self.standard_plates.value,
            standard_vol=self.standard_vol.value
        )
        return standard_dict


class Outputs:
    def __init__(self):
        self.style = Headers().style
        self.out_display = ipw.Output()
        self.output_dict = {}
        self.file = 'captured_data.json'
        self.inputs = None
        self.plate_display = None
        self.assay_display = None
        self.db_display = None
        self.standard_display = None
        self.reagent_display = None

        self.show_outputs_button = ipw.Button(
            description='Display Outputs',
            button_style='info'
        )
        self.show_outputs_button.on_click(self.run_outputs)
        display(self.show_outputs_button, self.out_display)

    def run_outputs(self, event):
        self.inputs = self.get_data_dict()
        self.plate_display, plate_details = TotalPlates(self.inputs).calculate_plates()
        self.output_dict.update(plate_details)

        if self.inputs["standard_plates"] > 0:
            self.standard_display, standard_data, standard_solution = StandardData(self.inputs).calculate_standards()
            self.output_dict['standard_solution'] = standard_solution
            self.output_dict['standard_data'] = standard_data
        else:
            self.standard_display = ipw.VBox([])

        self.assay_display, assay_details = Assays().calculate_assay(self.output_dict)
        self.output_dict.update(assay_details)

        self.db_display, db_details = DilutionBuffer().calculate_vols(self.inputs, self.output_dict)
        self.output_dict.update(db_details)

        # Temp use of excel data - no inventory
        self.reagent_display, reagent_details = ExcelReagents(self.inputs, self.output_dict).calculate_data()
        self.output_dict['reagent_details'] = reagent_details
        self.output_dict['assays'] = reagent_details
        calced_vols, folds = VolumeCalculations(self.inputs).calculate_volumes()
        self.output_dict['calced_vols'] = calced_vols
        self.output_dict['folds'] = folds

        if self.inputs['proj_type'] == 'Fermentation':
            pd_vols_data = VolumeCalculations(self.inputs).calculate_pd_volumes()
            self.output_dict['pd_vols_data'] = pd_vols_data

        self.output_section()

    def get_data_dict(self):
        try:
            with open(self.file) as captured_data:
                input_data = json.load(captured_data)
        except FileNotFoundError:
            return None
        else:
            return input_data

    def output_section(self):
        templatebuild_button = ipw.Button(description="Template", button_style='success')
        templatebuild_button.on_click(self.output_to_template)
        with self.out_display:
            self.out_display.clear_output()
            output_form = ipw.VBox([
                ipw.HTML("<h2>Output Section</h2>"),
                self.plate_display,
                self.assay_display,
                self.db_display,
                self.standard_display,
                self.reagent_display,
                templatebuild_button
            ])
            display(output_form)

    def output_to_template(self, event):
        TemplateBuilder(self.inputs, self.output_dict)
        display(ipw.HTML("<b>Data output to template!</b>"))


class TotalPlates:
    def __init__(self, input_dict):
        self.inputs = input_dict
        self.style = Headers().style
        self.total_greiner = ipw.IntText(
            description="Total 384w Greiner: ",
            style=self.style,
            disabled=True
        )
        self.total_proxiplates = ipw.IntText(
            description="Total ProxiPlates: ",
            style=self.style,
            disabled=True
        )

        self.total_pd = ipw.IntText(
            description="Total 96w Predilution: ",
            style=self.style,
            disabled=True
        )

    def calculate_plates(self):
        source = self.inputs["source"]
        replicates = self.inputs["replicates"]
        proj_type = self.inputs["proj_type"]

        self.total_pd.value = self.inputs["pd"]

        if proj_type == 'Fermentation':
            total_proxiplates = (source * 2) * (1 + self.total_pd.value)

        else:
            total_proxiplates = source * (1 + self.total_pd.value)

        self.total_greiner.value = total_proxiplates
        self.total_pd.value *= source

        if replicates == 'n + 2':
            if total_proxiplates == 1:
                total_proxiplates += 1
            else:
                total_proxiplates += 2
        elif replicates == 'n + 1':
            total_proxiplates += 1
        elif replicates == 'n * 2':
            total_proxiplates *= 2

        self.total_proxiplates.value = total_proxiplates
        display_form = self.setup_form()
        plate_details = self.plate_details()
        return display_form, plate_details

    def setup_form(self):
        plate_display = ipw.VBox([
            ipw.HTML("<h5><b>Total Plates</b></h5>"),
            ipw.HBox([
                self.total_proxiplates,
                self.total_greiner,
                self.total_pd
            ])
        ])
        return plate_display

    def plate_details(self):
        plate_details = dict(
            total_pd=self.total_pd.value,
            total_greiner=self.total_greiner.value,
            total_proxiplates=self.total_proxiplates.value
        )
        return plate_details


class Assays:
    def __init__(self):
        self.style = Headers().style
        self.proxi_wells = FixedHiPrBindCalcs().proxi_wells
        self.proxi_vols = FixedHiPrBindCalcs().proxi_well_vol
        self.ml_ul_conv = FixedHiPrBindCalcs().ml_ul_conv
        self.tempest_1 = FixedHiPrBindCalcs().tempest_comp_one
        self.tempest_2 = FixedHiPrBindCalcs().tempest_comp_two
        self.tempest_3 = FixedHiPrBindCalcs().tempest_comp_three
        self.assay_rxn = ipw.FloatText(description='AS for rxn (mL): ', style=self.style, disabled=True)
        self.assay_dead = ipw.FloatText(description='AS dead vol (mL): ', style=self.style, disabled=True)
        self.assay_req = ipw.FloatText(description='AS needed (mL): ', style=self.style, disabled=True)

    def calculate_assay(self, output_dict):
        try:
            proxiplates = output_dict['total_proxiplates']
        except KeyError:
            pass
        else:
            self.assay_rxn.value = (self.proxi_vols * self.proxi_wells * proxiplates) / self.ml_ul_conv
            self.assay_dead.value = round((self.proxi_vols * proxiplates * self.tempest_1 * self.tempest_2) /\
                                    self.ml_ul_conv + self.tempest_3, 3)
            self.assay_req.value = int((self.assay_rxn.value + self.assay_dead.value)) + 1
            display_form = self.setup_form()
            assay_details = self.capture_outputs()
            return display_form, assay_details

    def setup_form(self):
        # Assay Solution Display
        display_form = ipw.VBox([
                ipw.HTML('<h5><b>Total Assay Solutions Needed:</b></h5>'),
                ipw.HBox([
                    self.assay_rxn,
                    self.assay_dead,
                    self.assay_req
                ])
            ])
        return display_form

    def capture_outputs(self):
        assay_details = dict(
            assay_rxn=self.assay_rxn.value,
            assay_dead=self.assay_dead.value,
            assay_req=self.assay_req.value
        )
        return assay_details


class DilutionBuffer:
    def __init__(self):
        self.style = Headers().style
        self.ml_ul_conv = FixedHiPrBindCalcs().ml_ul_conv
        self.source_wells = FixedHiPrBindCalcs().source_wells
        self.proxi_wells = FixedHiPrBindCalcs().proxi_wells
        self.dbi_vol_total = ipw.IntText(description="DBI (mL): ", style=self.style, disabled=True)
        self.dbii_vol_total = ipw.IntText(description="DBII (mL): ", style=self.style, disabled=True)

    def calculate_vols(self, input_dict, output_dict):
        source = input_dict['source']
        pd_vols = input_dict['pd_vols']
        dbi_vol = input_dict['dbi_vol']
        dbii_vol = input_dict['dbii_vol']
        greiner = output_dict['total_greiner']
        total_pd_vol = sum([vol for vol in pd_vols.values()])
        # temp
        standard_scheme_total_dbi = 0

        pd_total_vol = -(-((total_pd_vol * int(source) * self.source_wells) / self.ml_ul_conv) // 1)

        self.dbi_vol_total.value = (int((int(source) * dbi_vol * self.source_wells) / self.ml_ul_conv) + 1)\
                                + pd_total_vol + standard_scheme_total_dbi
        self.dbii_vol_total.value = int((int(greiner) * dbii_vol * self.proxi_wells) / self.ml_ul_conv) + 1
        display_form = self.setup_form()
        db_details = self.capture_outputs()
        return display_form, db_details

    def setup_form(self):
        display_form = ipw.VBox([
                ipw.HTML('<h5><b>Total Dilution Buffers needed:</b></h5>'),
                ipw.HBox([
                    self.dbi_vol_total,
                    self.dbii_vol_total
                ])
            ])
        return display_form

    def capture_outputs(self):
        db_details = dict(
            dbi_vol_total=self.dbi_vol_total.value,
            dbii_vol_total=self.dbii_vol_total.value
        )
        return db_details


class StandardData:
    def __init__(self, input_data):
        self.style = Headers().style
        self.standard_buffer_amount = FixedHiPrBindCalcs().standard_buffer_amount
        self.ml_ul_conv = FixedHiPrBindCalcs().ml_ul_conv
        self.standard_total_vol = ipw.IntText(description="Total Standard Volume (uL):", style=self.style, disabled=True)
        self.standard_total_stock = ipw.FloatText(description="Total Stock Needed (uL): ", style=self.style, disabled=True)
        self.standard_total_dbi = ipw.FloatText(description="Total DBI needed (uL): ", style=self.style, disabled=True)
        self.proj_id = input_data["project_name_id"]
        self.project = input_data["project"]
        self.project_scheme = input_data["project_scheme"]
        self.standard_plates = input_data["standard_plates"]
        self.standard_wells = input_data["standard_wells"]
        self.standard_vol = input_data["standard_vol"]
        self.all_folds = input_data["standard_folds"]
        self.fold_1 = input_data["standard_folds"]["standard_fold_1"]
        self.conc_1 = input_data["standard_concs"]["standard_conc_1"]
        self.standard_stock_conc = input_data["standard_stock"]["standard_stock_conc"]
        self.standard_stock_mw = input_data["standard_stock"]["standard_stock_mw"]
        self.standard_base_wvol = 0

    def calculate_standards(self):
        self.standard_base_wvol = self.standard_plates * self.standard_wells * self.standard_vol + self.standard_buffer_amount
        self.standard_total_vol.value = self.standard_base_wvol * (1 / (self.fold_1 - 1) + 1)
        standard_stock_nm = (self.standard_stock_conc / self.standard_stock_mw) * 1000000000
        self.standard_total_stock.value = round((self.conc_1 * self.standard_total_vol.value) / standard_stock_nm, 2)
        self.standard_total_dbi.value = round(self.standard_total_vol.value - self.standard_total_stock.value, 2)
        standard_scheme_total_dbi = (self.standard_total_dbi.value + self.standard_base_wvol * 5) / self.ml_ul_conv

        standard_details = dict(
            standard_base_wvol=self.standard_base_wvol,
            standard_total_vol=self.standard_total_vol,
            standard_stock_conc=standard_stock_nm,
            standard_total_stock=self.standard_total_stock.value,
            standard_total_dbi=self.standard_total_dbi.value,
            standard_scheme_total_dbi=standard_scheme_total_dbi,
        )

        display_form, standard_data = self.setup_form()
        standard_solution = self.capture_outputs()
        return display_form, standard_data, standard_solution

    def setup_form(self):
        well_labels = [f"Well {num}" for num in range(2, 7)]
        transfer_vols = self.get_transfer_values()
        dbi_pre_vols = self.get_dbi_pre_vols()
        well_labels_display = self.create_display_box("", well_labels)
        dbi_vals_display = self.create_display_box("Add DBI (uL): ", dbi_pre_vols)
        transfer_vols_display = self.create_display_box("Transfer (uL): ", transfer_vols)
        standard_scheme_display = ipw.VBox([
            ipw.HTML(value="<h3>Standard Curve Prep</h3>"),
            ipw.HTML(value="<h5>Standard Stock Prep</h5>"),
            ipw.HBox([
                self.standard_total_vol,
                self.standard_total_stock,
                self.standard_total_dbi
            ]),
            ipw.HTML(value='<h5><b>Standard Dilution Scheme</b></h5>'),
            ipw.HBox([
                ipw.VBox([
                    ipw.Text(
                        value="Well 1",
                        style=self.style,
                        layout=ipw.Layout(width="auto")),
                    ipw.Text(
                        value=str(self.standard_total_vol.value),
                        style=self.style,
                        layout=ipw.Layout(width="auto")),
                    ipw.Label("")
                ]),
                ipw.VBox([
                    well_labels_display,
                    dbi_vals_display,
                    transfer_vols_display
                ])
            ])
        ])
        standard_data = [well_labels, dbi_pre_vols, transfer_vols]
        return standard_scheme_display, standard_data

    def get_transfer_values(self):
        transfer_vols = [str(round(self.standard_base_wvol / (fold - 1), 0)) for fold in self.all_folds.values()]
        return transfer_vols

    def get_dbi_pre_vols(self):
        dbi_pre_vols = [self.standard_base_wvol for repeat in range(0, 5)]
        return dbi_pre_vols

    def create_display_box(self, label, value_list):
        standard_display_boxwidth = '12%'
        display_box = ipw.HBox([ipw.Label(value=label,
                                          style=self.style,
                                          layout=ipw.Layout(
                                                 width=standard_display_boxwidth,
                                                 display='flex',
                                                 justify_content='flex-end'
                                             )),
                               ipw.HBox(
                                   [ipw.Text(value=str(value),
                                             layout=ipw.Layout(width=standard_display_boxwidth))
                                    for value in value_list])])
        return display_box

    def capture_outputs(self):
        standard_solution = dict(
            standard_total_vol=self.standard_total_vol.value,
            standard_total_stock=self.standard_total_stock.value,
            standard_total_dbi=self.standard_total_dbi.value
        )
        return standard_solution


class AssaySolutions:
    def __init__(self, input_dict, output_dict):
        self.inputs = input_dict
        self.outputs = output_dict
        self.style = Headers().style
        self.assay_req = output_dict['assay_req']
        self.proj_id = input_dict['project_name_id']

    def get_reagent_details(self):
        select_query = f"""
            SELECT 
                reagents.reagent_id, 
                project_reagents.assay_id, 
                reagent, 
                concentration_ugul, 
                on_hand, 
                project_reagents.desired_conc,
            CASE 
                WHEN reagent ILIKE 'lysozyme%' OR reagent ILIKE 'picogreen%'  
                THEN ROUND(on_hand - {self.assay_req} * project_reagents.desired_conc / concentration_ugul * 1000, 2)
                WHEN project_reagents.assay_id = 2 
                THEN ROUND(on_hand - (project_reagents.desired_conc  / concentration_ugul) * 1000 * {self.assay_req}, 2) 
                WHEN project_reagents.assay_id = 1 
                THEN ROUND(on_hand - (project_reagents.desired_conc * {self.assay_req} * 1000) / concentration_nm, 2)
                END as ul_remaining,
            CASE 
                WHEN reagent ILIKE 'lysozyme%' OR reagent ILIKE 'picogreen%' 
                THEN ROUND({self.assay_req} * project_reagents.desired_conc / concentration_ugul * 1000, 2)
                WHEN project_reagents.assay_id = 2 
                THEN ROUND((project_reagents.desired_conc  / concentration_ugul) * 1000 * {self.assay_req}, 2) 
                WHEN project_reagents.assay_id = 1 
                THEN ROUND((project_reagents.desired_conc *{self.assay_req} * 1000) / concentration_nm, 2)
                END as ul_needed
            FROM project_reagents
            
            INNER JOIN reagents 
            ON project_reagents.reagent_id = reagents.reagent_id
            WHERE project_reagents.proj_id = {self.proj_id}
            ORDER BY project_reagents.assay_id
        """
        reagent_details = self.db.query_call(select_query)
        reagent_details.insert(0,
                               ('ID', 'Assay', 'Reagent',
                                'Conc ug/ul', 'On Hand', 'Desired conc nM',
                                'Remaining uL', 'Needed ul')
                               )

        display_form = self.setup_form(reagent_details)
        reagent_tables = self.prepare_reagent_tables(reagent_details)
        # return display_form, reagent_details
        return display_form, reagent_tables

    def setup_form(self, data):
        display_form = ipw.VBox([
            ipw.HTML("<h5><b>Reagent Details</b></h5>"),
            ipw.VBox([
                ipw.HBox([
                    ipw.HTML(value=f'<b>{str(row_item[i])}</b>',
                             layout=ipw.Layout(
                                 width='12%',
                                 border='solid'),
                             style=self.style,
                             disabled=True)
                    if row_item[0] == 'ID' else
                    ipw.Text(value=str(row_item[i]),
                             layout=ipw.Layout(
                                 width='12%',
                                 border='0.5px solid'),
                             style=self.style,
                             disabled=True)
                    for i in range(0, len(row_item))
                 ]) for row_item in data])
            ])

        return display_form

    def prepare_reagent_tables(self, reagent_records):
        assay_list = list(set([value[1] for value in reagent_records if isinstance(value[1], int)]))
        assay_table_dict = {}
        for assay in assay_list:
            assay_table_dict[f"Assay {assay}"] = [row for row in reagent_records[1:] if row[1] == assay]
            assay_table_dict[f"Assay {assay}"].insert(0, reagent_records[0])
            assay_totals = ["", "", "", "", "", "", "Total Assay (mL):", self.assay_req * 1000]
            dbii_totals = ["", "", "", "", "", "", f"DBII for assay {assay} (mL):", round(self.assay_req - (sum(
                [float(row[-1]) for row in reagent_records[1:] if row[1] == assay]
            ) / 1000), 3) * 1000]
            assay_table_dict[f"Assay {assay}"].extend([dbii_totals])
            assay_table_dict[f"Assay {assay}"].extend([assay_totals])

        return assay_table_dict


class VolumeCalculations:
    def __init__(self, input_dict):
        self.dil_vols = input_dict['dil_vols']
        self.points = input_dict['points']
        self.cell_resus = input_dict['cell_resus']
        self.dbi_vol = input_dict['dbi_vol']
        self.dbii_vol = input_dict['dbii_vol']
        self.pd_vols = input_dict['pd_vols']
        self.pd_spikes = input_dict['pd_spikes']
        self.total_pd = input_dict['pd']

    def calculate_volumes(self):
        calced_vols = []
        folds = []
        pellet_conc = self.dbi_vol / self.cell_resus
        if self.points == 4:
            dil_vol_list = [dil_vol for dil_vol in self.dil_vols.values()]
        elif self.points == 8:
            dil_vol_list = [dil_vol for dil_vol in self.dil_vols.values()] * 2
        for dil_vol in dil_vol_list:
            dil_factor = 1 + self.dbii_vol / dil_vol
            fold_factor = (pellet_conc * dil_factor)
            folds.append(fold_factor)
            calced_vol = 6 / fold_factor
            calced_vols.append(calced_vol)
            pellet_conc = fold_factor

        return calced_vols, folds

    def calculate_pd_volumes(self):
        pd_data = {}
        initial_cell_conc = self.dbi_vol / self.cell_resus

        for plate in range(1, self.total_pd + 1):
            initial_pd_conc = self.pd_vols[f'pd_{plate}_vol'] / self.pd_spikes[f'pd_{plate}_spike'] + 1
            initial_conc = initial_cell_conc * initial_pd_conc
            calced_vols = []
            folds = []
            dil_vol_list = [dil_vol for dil_vol in self.dil_vols.values()] * 2
            for dil_vol in dil_vol_list:
                dil_factor = 1 + self.dbii_vol / dil_vol
                fold_factor = (initial_conc * dil_factor)
                folds.append(fold_factor)
                calced_vol = 6 / fold_factor
                calced_vols.append(calced_vol)
                initial_conc = fold_factor
            pd_data[f'pd_{plate}'] = dict(
                pd_folds=folds,
                pd_calced_vols=calced_vols
            )
        return pd_data


class ExcelReagents:
    def __init__(self, input_dict, output_dict):
        self.input_dict = input_dict
        self.output_dict = output_dict
        self.project = input_dict['project']
        self.scheme = input_dict['project_scheme']
        self.style = Headers().style
        self.reagents = FixedHiPrBindCalcs().excel_reagents
        self.reagent_dict = self.get_reagents()

    def get_reagents(self):
        reagent_data = ExcelData().get_reagent_data(self.project, self.scheme)
        reagent_dict = reagent_data.to_dict(orient='records')

        return reagent_dict

    def calculate_data(self):
        assay_req = self.output_dict['assay_req']
        reagent_records = [["Reagent", "Assay", "Cat_Num", "Conc. (ug/uL)", 'Desired conc. (nM)', 'Needed vol (uL)']]

        for reagent in self.reagents:
            conc_ug_ul = reagent[-1]
            desired_conc = reagent[-2]
            needed_vol = round(assay_req * desired_conc / conc_ug_ul * 1000, 2)
            reagent.append(needed_vol)
            reagent_records.append(reagent)

        for reagent in self.reagent_dict:
            try:
                int(reagent["Conc. (ug/uL)"])
            except ValueError:
                pass
            else:
                reagent_name = reagent["Reagent"]
                assay_num = int(reagent["Assay"])
                desired_conc = reagent['Desired conc. (nM)']
                cat_num = reagent["Cat. No./Code"]

                if 'bead' in reagent["Reagent"]:
                    conc_ug_ul = reagent["Conc. (ug/uL)"]
                    needed_vol = round((desired_conc / conc_ug_ul) * 1000 * assay_req, 2)
                    reagent_record = [reagent_name, assay_num, cat_num, conc_ug_ul, desired_conc, needed_vol]

                else:
                    conc_nm = reagent['Conc. (nM)']
                    needed_vol = round(desired_conc * (assay_req * 1000 / conc_nm), 2)
                    reagent_record = [reagent_name, assay_num, cat_num, conc_nm, desired_conc, needed_vol]

                reagent_records.append(reagent_record)

        reagent_display = self.setup_form(reagent_records)
        assay_details = self.prepare_reagent_tables(reagent_records, assay_req)
        return reagent_display, assay_details

    def setup_form(self, data):
        display_form = ipw.VBox([
            ipw.HTML("<h5><b>Reagent Details</b></h5>"),
            ipw.VBox([
                ipw.HBox([
                    ipw.HTML(value=f'<b>{str(row_item[i])}</b>',
                             layout=ipw.Layout(
                                 width='12%',
                                 border='solid'),
                             style=self.style,
                             disabled=True)
                    if row_item[0] == 'Reagent' else
                    ipw.Text(value=str(row_item[i]),
                             layout=ipw.Layout(
                                 width='12%',
                                 border='0.5px solid'),
                             style=self.style,
                             disabled=True)
                    for i in range(0, len(row_item))
                 ]) for row_item in data])
            ])

        return display_form

    def prepare_reagent_tables(self, reagent_records, assay_req):
        assay_list = list(
            set(
                [int(reagent["Assay"]) for reagent in self.reagent_dict if not isinstance(reagent["Assay"], dict)]
            )
        )
        assay_db = {}
        assay_table_dict = {}
        for assay in assay_list:
            assay_table_dict[f"Assay {assay}"] = [row for row in reagent_records[1:] if row[1] == assay]
            assay_table_dict[f"Assay {assay}"].insert(0, reagent_records[0])
            assay_totals = ["", "", "", "", "Total Assay (mL):", assay_req * 1000]
            dbii_totals = ["", "", "", "", f"DBII for assay {assay} (mL):", round(assay_req - (sum(
                [float(row[-1]) for row in reagent_records[1:] if row[1] == assay]
            ) / 1000), 3) * 1000]
            assay_table_dict[f"Assay {assay}"].extend([dbii_totals])
            assay_table_dict[f"Assay {assay}"].extend([assay_totals])
        return assay_table_dict


class TemplateBuilder:
    def __init__(self, input_dict, output_dict):
        self.input_dict = input_dict
        self.output_dict = output_dict
        self.project_name = input_dict["project"]
        self.proj_id = input_dict["proj_id"]
        self.proj_file_option = input_dict["proj_file_option"]
        self.proj_type = input_dict["proj_type"]
        self.project_scheme = input_dict["project_scheme"]
        self.output_folder = r"outputs/"
        self.ssf_path = r"outputs\SSF-HPB-runs"
        self.ferm_path = r"outputs\Ferm-HPB-runs"
        self.parent_directory = f"{self.proj_id}_{self.project_name}_{self.proj_type}"
        self.sub_directory = "Analysis"
        self.filename = self.make_outfile()
        self.output_file = self.make_folders()
        self.template = Templates(self.project_name, self.proj_type).fetch_template()

        self.write_to_excel()

    def make_outfile(self):
        if self.proj_file_option:
            filename = f"{self.proj_id}_protocol_{self.project_scheme}_{self.proj_file_option}.xlsx"
        else:
            filename = f"{self.proj_id}_protocol_{self.project_scheme}.xlsx"
        return filename

    def make_folders(self):
        if self.proj_type == "Fermentation":
            dir_path = os.path.join(self.ferm_path, self.parent_directory)
        else:
            dir_path = os.path.join(self.ssf_path, self.parent_directory)

        output_file = os.path.join(dir_path, self.filename)
        if not os.path.exists(dir_path):
            os.mkdir(dir_path)
            analysis_path = os.path.join(dir_path, self.sub_directory)
            os.mkdir(analysis_path)

        return output_file

    def write_to_excel(self):
        location_dict = self.template['location_dict']
        wb = self.template['workbook']
        ws = self.template['worksheet']
        for in_key, in_value in self.input_dict.items():
            if isinstance(in_value, dict):
                for sub_key, sub_value in in_value.items():
                    if sub_key in location_dict:
                        ws[location_dict[sub_key]].value = sub_value
            elif in_key in location_dict:
                ws[location_dict[in_key]].value = in_value
            else:
                pass
        for in_key, in_value in self.output_dict.items():
            if in_key == "pd_vols_data":
                start_row = location_dict[in_key][0]
                data_start_row = start_row
                for plate, plate_data in in_value.items():
                    start_col = location_dict[in_key][1]
                    for inner_data, inner_list in plate_data.items():
                        data_start_row = start_row
                        for value in inner_list:
                            ws.cell(row=data_start_row, column=start_col).value = value
                            data_start_row += 1
                        start_col += 1
                    start_row = data_start_row + 3

            elif isinstance(in_value, list) and in_key in location_dict:
                start_row = location_dict[in_key][0]
                for row in in_value:
                    start_col = location_dict[in_key][1]
                    if isinstance(row, list) or isinstance(row, tuple):
                        for value in row:
                            ws.cell(row=start_row, column=start_col).value = value
                            start_col += 1
                    else:
                        ws.cell(row=start_row, column=start_col).value = row
                    start_row += 1
            elif isinstance(in_value, dict) and in_key in location_dict:
                start_row = location_dict[in_key][0]
                start_col = location_dict[in_key][1]
                for in_dict_key, in_dict_list in in_value.items():
                    ws.cell(row=start_row, column=start_col).value = in_dict_key
                    start_row += 1
                    for row in in_dict_list:
                        for value in row:
                            ws.cell(row=start_row, column=start_col).value = value
                            start_col += 1
                        start_col = location_dict[in_key][1]
                        start_row += 1
                    start_row += 1

            elif isinstance(in_value, dict):
                for sub_key, sub_value in in_value.items():
                    if sub_key in location_dict:
                        ws[location_dict[sub_key]].value = sub_value
            elif in_key in location_dict:
                ws[location_dict[in_key]].value = in_value

            else:
                pass

        wb.save(self.output_file)


class Templates:
    def __init__(self, project, proj_type):
        self.project_name = project
        self.proj_type = proj_type
        self.template_dict = {}

    def fetch_template(self):
        if self.proj_type == "Fermentation":
            location_dict = dict(
                project="B1",
                proj_id="B2",
                proj_type="B3",
                run_notes="A5",
                source="A20",
                total_greiner="A23",
                total_proxiplates="A24",
                pd="A21",
                total_pd="A22",
                cell_resus="C42",
                dbi_vol="H42",
                dbii_vol="E52",
                dbi_vol_total="F20",
                dbii_vol_total="F22",
                dil_vol_1="D52",
                dil_vol_2="D53",
                dil_vol_3="D54",
                dil_vol_4="D55",
                pd_1_spike="C43",
                pd_2_spike="C44",
                pd_3_spike="C45",
                pd_4_spike="C46",
                pd_1_vol="H43",
                pd_2_vol="H44",
                pd_3_vol="H45",
                pd_4_vol="H46",
                standard_total_vol="F34",
                standard_total_stock="F33",
                standard_total_dbi="F32",
                standard_stock_conc="C30",
                standard_stock_mw="E30",
                standard_conc_1="B32",
                standard_conc_2="B33",
                standard_conc_3="B34",
                standard_conc_4="B35",
                standard_conc_5="B36",
                standard_conc_6="B37",
                standard_data=(35, 9),
                calced_vols=(52, 10),
                folds=(52, 9),
                assays=(62, 2),
                pd_vols_data=(52, 13)
            )
            wb = load_workbook(r"templates/Standard Protocol_FER_tabs.xlsx")
        else:
            location_dict = dict(
                project="B1",
                proj_id="B2",
                proj_type="B3",
                run_notes="A5",
                source="B19",
                total_greiner="B20",
                total_proxiplates="B21",
                cell_resus="C36",
                dbi_vol="G36",
                dbii_vol="E39",
                dbi_vol_total="J19",
                dbii_vol_total="J21",
                dil_vol_1="D39",
                dil_vol_2="D40",
                dil_vol_3="D41",
                dil_vol_4="D42",
                standard_total_vol="F28",
                standard_total_stock="F27",
                standard_total_dbi="F26",
                standard_stock_conc="C24",
                standard_stock_mw="E24",
                standard_conc_1="B26",
                standard_conc_2="B27",
                standard_conc_3="B28",
                standard_conc_4="B29",
                standard_conc_5="B30",
                standard_conc_6="B31",
                standard_data=(29, 9),
                calced_vols=(39, 11),
                folds=(39, 10),
                assays=(45, 2),
            )
            self.template_dict['location_dict'] = location_dict
            wb = load_workbook(r"templates/Standard Protocol_SSF_tabs.xlsx")

        self.template_dict['location_dict'] = location_dict
        self.template_dict['workbook'] = wb
        ws = wb['Run Info']
        self.template_dict['worksheet'] = ws

        return self.template_dict
